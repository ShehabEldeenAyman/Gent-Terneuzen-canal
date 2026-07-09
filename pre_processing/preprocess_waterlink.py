"""
Cleans the 'result' tab of data.xlsx.

Assumed raw layout (1-indexed rows/cols, openpyxl convention):
    Row 1        -> free-text description        (col A = None, col B = None, col C.. = description or None)
    Row 2        -> attribute / tag name          (col A = 'Datum', col B = 'Tijd', col C.. = tag name)
    Row 3        -> unit of measure               (col A = 'eenheid', col B = None, col C.. = unit or None)
    Row 4..end   -> data. Col A = date (datetime), Col B = time (string 'HH:MM:SS')

Output layout:
    Row 1        -> single combined attribute header per column
                    ('DateTime' for the merged date/time column,
                     'description | tag_name | unit' for every measurement column,
                     any part that is missing/None is simply omitted)
    Row 2..end   -> pure data instances, with columns A+B of the original replaced
                    by a single ISO-8601 datetime string, e.g. '2021-03-03T23:45:00+00:00'
"""

import csv
from datetime import datetime, date, time as time_cls
import openpyxl


def combine_datetime(date_val, time_val, tz_offset: str = "+00:00") -> str:
    """
    Combine a date value and a time value into a single ISO-8601 string,
    e.g. '2021-03-03T23:45:00+00:00'.

    date_val : datetime.datetime / datetime.date / str
    time_val : datetime.time / str ('HH:MM:SS') / None
    tz_offset: string appended at the end (default UTC, '+00:00')
    """
    # --- normalize the date part ---
    if isinstance(date_val, datetime):
        d = date_val.date()
    elif isinstance(date_val, date):
        d = date_val
    elif isinstance(date_val, str):
        d = datetime.fromisoformat(date_val.strip()).date()
    else:
        raise TypeError(f"Unsupported date type: {type(date_val)}")

    # --- normalize the time part ---
    if time_val is None:
        t = time_cls(0, 0, 0)
    elif isinstance(time_val, time_cls):
        t = time_val
    elif isinstance(time_val, datetime):
        t = time_val.time()
    elif isinstance(time_val, str):
        t = datetime.strptime(time_val.strip(), "%H:%M:%S").time()
    else:
        raise TypeError(f"Unsupported time type: {type(time_val)}")

    combined = datetime.combine(d, t)
    return combined.strftime("%Y-%m-%dT%H:%M:%S") + tz_offset


def build_combined_header(description, name, unit, separator: str = " | ") -> str:
    """
    Combine the description / attribute name / unit of measure (rows 1, 2, 3)
    into a single attribute label for row 1 of the cleaned sheet.
    None / empty parts are skipped.
    """
    parts = [str(p).strip() for p in (description, name, unit) if p not in (None, "")]
    return separator.join(parts) if parts else ""


def clean_result_sheet(
    input_path: str,
    output_path: str,
    sheet_name: str = "result",
    date_col: int = 1,       # column A
    time_col: int = 2,       # column B
    first_data_row: int = 4, # row where actual observations start
    tz_offset: str = "+00:00",
    datetime_header: str = "DateTime",
):
    """
    Reads `sheet_name` from `input_path`, merges the Datum/Tijd columns into
    a single ISO-8601 datetime column, consolidates the 3-row header into a
    single row-1 header, and writes the result as a CSV to `output_path`.
    """
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    ws_in = wb_in[sheet_name]

    max_row = ws_in.max_row
    max_col = ws_in.max_column

    # --- read the 3 header rows ---
    row_desc = [ws_in.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    row_name = [ws_in.cell(row=2, column=c).value for c in range(1, max_col + 1)]
    row_unit = [ws_in.cell(row=3, column=c).value for c in range(1, max_col + 1)]

    # --- build the new combined header ---
    # data columns are everything except date_col/time_col
    data_cols = [c for c in range(1, max_col + 1) if c not in (date_col, time_col)]

    new_header = [datetime_header]
    for c in data_cols:
        idx = c - 1
        new_header.append(
            build_combined_header(row_desc[idx], row_name[idx], row_unit[idx])
        )

    # --- write CSV, streaming row by row ---
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(new_header)

        for r in range(first_data_row, max_row + 1):
            date_val = ws_in.cell(row=r, column=date_col).value
            time_val = ws_in.cell(row=r, column=time_col).value

            if date_val is None:
                continue  # skip fully empty trailing rows

            combined_dt = combine_datetime(date_val, time_val, tz_offset=tz_offset)

            row_values = [combined_dt]
            for c in data_cols:
                row_values.append(ws_in.cell(row=r, column=c).value)

            writer.writerow(row_values)

    return output_path


if __name__ == "__main__":
    clean_result_sheet(
        input_path="/mnt/user-data/uploads/data.xlsx",
        output_path="/mnt/user-data/outputs/data_cleaned.csv",
        sheet_name="result",
    )
    print("Done.")